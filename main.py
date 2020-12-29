import json
from datetime import datetime

from loguru import logger
from openpyxl import load_workbook
import lxml.html
import requests
import re
from PIL import Image
import urllib.request
import io
import sys
import yaml
import random as r

from openpyxl.styles import PatternFill
from urllib3.exceptions import InsecureRequestWarning
from py_files.locators import LabirintLocators, BooksLocators, OzonLocators

from py_files.logging import set_logger, my_exception_hook


def user_agents_unpack(path_to_file):
    with open(path_to_file, encoding='utf-8') as user_agents_file:
        user_agents_list = yaml.load(user_agents_file, Loader=yaml.FullLoader)
        return user_agents_list


def get_input_data(row_data):
    number = row_data[0].value
    isbn = row_data[1].value
    keeping = row_data[2].value
    lang = row_data[3].value
    barcode = row_data[4].value

    if not keeping:
        keeping = 'Отличная'
    elif keeping == 1:
        keeping = 'Хорошая'

    if not lang:
        lang = 'Русский'
    elif lang == 1:
        lang = 'Английский'

    return number, isbn, keeping, lang, barcode


def filing_empty_sheet(rows_count, sheet):
    rows_count += 1
    sheet = work_book[sheet]
    cell_range = sheet[f'A{rows_count}':f'D{rows_count}']
    data_list = [NUMBER, ISBN, KEEPING, LANG, BARCODE]

    for each_row in cell_range:
        for each_cell in each_row:
            text_in_cell = str(data_list[0])
            each_cell.value = text_in_cell
            data_list.pop(0)
    work_book.save('Books_info_1.xlsx')
    return rows_count


def check_html_element_existing(element_list):
    if element_list:
        element_value = element_list[0]
        # if element_value.isdigit():
        #     element_value = int(element_value)
    else:
        element_value = ''
    return element_value


def check_for_book_existing(isbn):
    search_book_by_ISBN_link = f'https://www.labirint.ru/search/{isbn}/?stype=0'
    html_text_search_page = requests.get(search_book_by_ISBN_link, verify=False).text
    tree = lxml.html.document_fromstring(html_text_search_page)
    book_href = tree.xpath(LabirintLocators.product_cover)
    if book_href:
        return True
    else:
        full_book_link = f'https://www.books.ru/search.php?s%5Btype_of_addon%5D=all&s%5Bquery%5D={isbn}&s%5Bgo%5D=1'
        book_page_response = requests.get(full_book_link).text
        book_page_tree = lxml.html.document_fromstring(book_page_response)
        sell_this_book_button = book_page_tree.xpath('//div[@class=" book-navigation_item"]')
        if not sell_this_book_button:
            return False
    return True


def get_link_to_ozon_page(isbn_code):
    search_book_by_ISBN_link = f'https://www.ozon.ru/search/?from_global=true&text={isbn_code}'
    logger.debug(f'Ссылка на страницу поиска книги: {search_book_by_ISBN_link}')
    html_text_from_ozon = requests.get(search_book_by_ISBN_link, headers=header, verify=False).text
    tree_ozon_search = lxml.html.document_fromstring(html_text_from_ozon)

    check_exist = tree_ozon_search.xpath(OzonLocators.isbn_exist)
    if check_exist:
        logger.debug(f'ISBN: {isbn_code}')

        try:
            book_cheapest_page_link_element = tree_ozon_search.xpath(OzonLocators.cheapest_book_page_link)
            if book_cheapest_page_link_element:
                book_ozon_page_link = 'https://www.ozon.ru' + \
                                      book_cheapest_page_link_element[0].attrib['href']
                return book_ozon_page_link
            elif tree_ozon_search.xpath(OzonLocators.product_page_link):
                book_ozon_page_link = 'https://www.ozon.ru' + \
                                      tree_ozon_search.xpath(OzonLocators.product_page_link)[0].attrib[
                                          'href']
                logger.debug(f'Ссылка на книгу Ozon: {book_ozon_page_link}')
                return book_ozon_page_link

        except IndexError:
            logger.critical('Либо эротика либо баг')
            return None
    else:
        return None


def get_ozon_tree(link):
    html_text = requests.get(link, verify=False).text
    # print(html_text)
    tree_page = lxml.html.document_fromstring(html_text)
    return tree_page


def get_price_from_ozon(page_tree):
    print(page_tree.xpath(OzonLocators.price))
    if page_tree.xpath(OzonLocators.price):
        weird_div = page_tree.xpath(OzonLocators.price)[0]
        print(weird_div)
        weird_div_json = json.loads(weird_div)
        price_from_ozon = weird_div_json['cellTrackingInfo']['product'].get('finalPrice')
        print(price_from_ozon)
        if not price_from_ozon:
            # price_from_ozon = weird_div_json['cellTrackingInfo']['product'].get('finalPrice')
            logger.critical('Не нашло цену озона в Json')
            return None
        else:
            if price_from_ozon > 202:
                return price_from_ozon
            else:
                return 187


# def get_premium_price_from_ozon(page_tree):
#     premium_price_list = page_tree.xpath(OzonLocators.price_ozon_premium)
#     delete_this_var = check_html_element_existing(premium_price_list)
#     delete_this_var = re.findall(r'\d', check_html_element_existing(premium_price_list))
#     premium_price = ''.join(re.findall(r'\d', check_html_element_existing(premium_price_list)))
#
#     return premium_price

def calculate_price_ozon_for_column_f(column_D_price, vol):
    if column_D_price <= 400:
        if vol < 1:
            if column_D_price < 168:
                return column_D_price - 21
            elif column_D_price > 168:
                return None
        if vol > 1:
            if column_D_price < 208:
                return None
            elif column_D_price > 208:
                return column_D_price - 21
    elif column_D_price > 400:
        return column_D_price * 0.94
    logger.critical('Что-то не так с калькуляцией цены Озон Премиум')


def get_photo_link_from_ozon(page_tree):
    photo_link_list = page_tree.xpath(OzonLocators.photo_link)
    if photo_link_list:
        photo_link = photo_link_list[0].attrib['src']
        return photo_link


def check_if_volume_is_more_than_one_and_more_than_one_and_a_half(length, width, height):
    volume = length * width * height / 1000000
    if volume < 1 or volume > 1.15:
        return True

    elif 1 < volume < 1.15:
        return False
    else:
        logger.critical('Произошло что-то странное с объемом книги')


def calculate_minimal_price(volume):
    if volume < 1:
        return 147
    elif volume > 1.15:
        return 187


def calculate_price_ozon_book_for_column_D(minimal_price, price_from_ozon):
    if price_from_ozon > minimal_price + 15:
        return price_from_ozon - 15
    else:
        return minimal_price


def calculate_price_labirint_book(minimal_price, price_from_labirint):
    if price_from_labirint * 0.6 > minimal_price:
        return price_from_labirint * 0.6
    else:
        return minimal_price


def calculate_old_price(ozon_price, old_price_from_labirint, price_without_discount_from_labirint):
    if old_price_from_labirint:
        return int(old_price_from_labirint)
    elif price_without_discount_from_labirint:
        return int(price_without_discount_from_labirint)
    elif ozon_price:
        return int(ozon_price * 1.7)
    elif not ozon_price and not old_price_from_labirint and not price_without_discount_from_labirint:
        return None
    else:
        logger.critical(f'Что-то не так с ценой: '
                        f'Цена до скидки: {old_price_from_labirint}, '
                        f'Цена с озона:{ozon_price}, '
                        f'Цена без скидки: {price_without_discount_from_labirint}')


def define_capture_size(url):
    with urllib.request.urlopen(url) as u:
        f = io.BytesIO(u.read())
        img = Image.open(f)

    picture_sizes = img.size

    for size in picture_sizes:
        if size < 600:
            return False

    return True


def get_book_link(isbn):
    search_book_by_ISBN_link = f'https://www.labirint.ru/search/{isbn}/?price_min=&price_max=&age_min=&age_max=&form-pubhouse=&lit=&stype=0&available=1&wait=1&no=1&preorder=1&paperbooks=1'
    html_text_labirint_search = requests.get(search_book_by_ISBN_link, verify=False).text
    tree = lxml.html.document_fromstring(html_text_labirint_search)

    book_href = tree.xpath(LabirintLocators.product_cover)
    if book_href:
        store_locators = LabirintLocators
        book_href = book_href[0]
        full_book_link = 'https://www.labirint.ru' + book_href
    else:
        store_locators = BooksLocators
        full_book_link = f'https://www.books.ru/search.php?s%5Btype_of_addon%5D=&s%5Bquery%5D={isbn}&s%5Bgo%5D=1'
    return store_locators, full_book_link


def get_data_from_genres_file(genres_file_name):
    genres_file = load_workbook(genres_file_name)
    sheet_genres = genres_file['Лист1']

    count = 0
    genres_dict = {}
    for cell in sheet_genres['B']:
        count += 1
        if cell.value:
            genres_dict[cell.value] = sheet_genres[f'C{count}'].value
        else:
            break

    return genres_dict


def get_years_sequence(start_year, stop_year):
    years_gap_list = []
    for i in range(start_year, stop_year):
        years_gap_list.append(str(i))
    return years_gap_list


def get_site_tree(link):
    html_text = requests.get(link, verify=False).text
    tree = lxml.html.document_fromstring(html_text)
    return tree


def define_which_sheet(years, tree):
    year_list = tree.xpath(site_locators.year_xpath)

    if year_list:
        year = check_html_element_existing(year_list)
        year = ''.join(re.findall(r'\d', year))

        if year in years:
            return 'Second-hand', int(year)

        else:
            return 'Букинистика', int(year)

    else:
        logger.critical('Нет года. Проверь, что все записалось в правильный лист')
        return 'Без информации', ''


def get_cover_from_labirint(link_book):
    book_id = link_book.split("/")[-2]

    design_link = f'https://www.labirint.ru/ajax/design/{book_id}/'

    headers = {
        'accept': '*/*',
        'user-agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/85.0.4183.83 Safari/537.36',
        'x-requested-with': 'XMLHttpRequest',
        'sec-fetch-site': 'same-origin',
        'sec-fetch-mode': 'cors',
        'sec-fetch-dest': 'empty',
        'referer': link_book

    }

    response_for_cover_html = requests.get(design_link, headers=headers).text
    tree_cover = lxml.html.document_fromstring(response_for_cover_html)

    cover_text = tree_cover.xpath(LabirintLocators.cover_xpath)
    cover_text = check_html_element_existing(cover_text)

    if 'мягкий' in cover_text:
        cover = 'Мягкая обложка'
    else:
        cover = 'Твердый переплет'

    design_list = tree_cover.xpath(LabirintLocators.colored_pics_xpath)
    design_text = ''.join(design_list)

    if design_text:
        if 'Цветные' in design_text:
            colored_pics = 'Да'
        else:
            colored_pics = ''
    else:
        colored_pics = ''

    return cover, colored_pics


def get_data(tree):
    if site_locators == LabirintLocators:

        circulation = ''

        name_list = tree.xpath(site_locators.name_xpath)
        name = check_html_element_existing(name_list)
        if name:
            if ':' in name:
                name = re.findall(r': (.*)', name)[0]

        logger.debug(f'Книга "{name}": {book_link}')

        cover, colored_pics = get_cover_from_labirint(book_link)

        genres_from_html = tree.xpath(LabirintLocators.genres_xpath)
        if genres_from_html:
            genres_from_html.reverse()
            for elem in genres_from_html:
                if elem in GENRES_LABIRINT_LIST.keys():
                    genre = GENRES_LABIRINT_LIST[elem]
                    break
                else:
                    genre = ''
        else:
            genre = ''

        annotation = tree.xpath(site_locators.annotation_xpath)
        if annotation:
            annotation = ''.join(annotation)
        else:
            annotation = ''

        no_price = tree.xpath(site_locators.no_price_xpath)
        if no_price:
            discount_price_from_labirint = ''
            old_price_with_discount_from_labirint = ''
            price_without_discount_from_labirint = ''

        else:
            price_without_discount_list = tree.xpath(site_locators.price_without_discount_xpath)

            if price_without_discount_list:
                price_without_discount_from_labirint = int(price_without_discount_list[0])
                discount_price_from_labirint = ''
                old_price_with_discount_from_labirint = ''

            else:
                old_price_with_discount_from_labirint = check_html_element_existing(
                    tree.xpath(site_locators.old_price_xpath))

                if old_price_with_discount_from_labirint:
                    old_price_with_discount_from_labirint = int(old_price_with_discount_from_labirint)

                discount_price_from_labirint = check_html_element_existing(tree.xpath(site_locators.new_price_xpath))

                if discount_price_from_labirint:
                    discount_price_from_labirint= int(discount_price_from_labirint)

                price_without_discount_from_labirint = ''

        orig_name_list = tree.xpath(site_locators.orig_name_xpath)
        orig_name = check_html_element_existing(orig_name_list)

        editor_list = tree.xpath(site_locators.editor_xpath)
        editor = check_html_element_existing(editor_list)

        illustrator_list = tree.xpath(site_locators.illustrator_xpath)
        illustrator = check_html_element_existing(illustrator_list)
        if illustrator:
            illustrator = ', '.join(illustrator_list)


    elif site_locators == BooksLocators:
        name_list = tree.xpath(site_locators.name_xpath)

        name = check_html_element_existing(name_list)

        logger.debug(f'Книга "{name}": {book_link}')

        editor = ''

        illustrator = ''

        orig_name = ''

        full_annotation = tree.xpath(site_locators.full_annotation_xpath)
        if full_annotation:
            annotation = ''.join(full_annotation)
        else:
            short_annotation = tree.xpath(site_locators.short_annotation_xpath)
            annotation = check_html_element_existing(short_annotation)

        colored_pics = ''

        circulations_list = tree.xpath(site_locators.circulation)
        circulation = check_html_element_existing(circulations_list)
        if circulation:
            circulation = ''.join(re.findall(r'\d', circulation))

        genre = check_html_element_existing(tree.xpath(BooksLocators.genres_xpath))

        # if genres_list:
        #     genres_list.reverse()
        #     for elem in genres_list:
        #         if elem in GENRES_BOOKS_RU_LIST.keys():
        #             genre = GENRES_BOOKS_RU_LIST[elem]
        #             break
        #         else:
        #             genre = ''
        # else:
        #     genre = ''

        discount_price_from_labirint = ''
        old_price_with_discount_from_labirint = ''
        price_without_discount_from_labirint = ''

        cover_list = tree.xpath(BooksLocators.cover_xpath)
        cover = check_html_element_existing(cover_list)

        if 'мягкий' in cover:
            cover = 'Мягкая обложка'
        else:
            cover = 'Твердый переплет'

    weight_list = tree.xpath(site_locators.weight_xpath)
    weight = check_html_element_existing(weight_list)
    if weight:
        weight = int(''.join(re.findall(r'\d', weight)))

    dimension_list = tree.xpath(site_locators.dimensions_xpath)

    dimensions = check_html_element_existing(dimension_list)
    if dimensions:
        if re.search(r'\d', dimensions):
            dimensions = re.split(r'\D', (''.join(re.sub(r' ', '', dimensions))))
            dimensions = list(filter(None, dimensions))

            for parameter in dimensions:
                dimensions[dimensions.index(parameter)] = int(parameter)

            if len(dimensions) == 1:
                length = int(dimensions[0])
                if site_locators == BooksLocators:
                    length = int(length * 10)
                width = ''
                height = ''

            elif len(dimensions) == 2:
                length = int(dimensions[0])
                width = int(dimensions[1])
                if site_locators == BooksLocators:
                    length = int(length * 10)
                    width = int(width * 10)
                height = ''

            elif len(dimensions) == 3:
                length = int(dimensions[0])
                width = int(dimensions[1])
                height = int(dimensions[2])
                if site_locators == BooksLocators:
                    length = length * 10
                    width = width * 10
                    height = height * 10

        else:
            width = ''
            height = ''
            length = ''

    else:
        width = ''
        height = ''
        length = ''

    author_code = tree.xpath(site_locators.author_xpath)
    author = ''
    for elem in author_code:
        if re.search('[а-яА-Я]', elem):
            author += elem.strip()
            break

    publisher_list = tree.xpath(site_locators.publisher_xpath)
    publisher = check_html_element_existing(publisher_list)

    series_list = tree.xpath(site_locators.series_xpath)
    series = check_html_element_existing(series_list)

    pages_list = tree.xpath(site_locators.pages_xpath)
    pages = check_html_element_existing(pages_list)
    if pages:
        pages = ''.join(re.findall(r'\d', pages))

    # TODO: посмотреть, можно ли пофиксить дублировать логирования ISBN Озон книг
    if not ozon_exist:
        logger.debug(f'ISBN: {ISBN}')

    data_dict = {'number': NUMBER, 'name': name, 'discount_price_from_labirint': discount_price_from_labirint,
                 'old_price_with_discount_from_labirint': old_price_with_discount_from_labirint,
                 'price_without_discount_from_labirint': price_without_discount_from_labirint, 'barcode': BARCODE,
                 'weight': weight, 'width': width, 'height': height, 'length': length, 'photo_link': photo_link,
                 'isbn': ISBN, 'genre': genre, 'author': author, 'annotation': annotation,
                 'publisher': publisher, 'year': YEAR, 'series': series, 'pages': pages,
                 'colored_pics': colored_pics, 'lang': LANG, 'orig_name': orig_name, 'keeping': KEEPING,
                 'editor': editor, 'illustrator': illustrator, 'circulation': circulation, 'cover': cover}

    return data_dict


def save_to_table(row_count, sheet, data_dict):
    sheet = work_book[sheet]
    sheet[f'A{row_count}'].value = data_dict['number']
    sheet[f'B{row_count}'].value = data_dict['barcode']
    sheet[f'C{row_count}'].value = data_dict['name']
    sheet[f'D{row_count}'].value = data_dict.get('column_D_price')
    sheet[f'E{row_count}'].value = data_dict.get('column_E_price')
    sheet[f'F{row_count}'].value = data_dict.get('column_F_ozon_price')
    sheet[f'J{row_count}'].value = data_dict['barcode']
    sheet[f'K{row_count}'].value = data_dict['weight']
    sheet[f'L{row_count}'].value = data_dict['width']
    sheet[f'M{row_count}'].value = data_dict['height']
    sheet[f'N{row_count}'].value = data_dict['length']
    sheet[f'O{row_count}'].value = data_dict['photo_link']
    sheet[f'T{row_count}'].value = 'Книга'
    sheet[f'U{row_count}'].value = data_dict['isbn']
    sheet[f'V{row_count}'].value = data_dict['genre']
    sheet[f'W{row_count}'].value = 'Печатная книга'
    sheet[f'X{row_count}'].value = data_dict['author']
    sheet[f'AA{row_count}'].value = data_dict['cover']
    sheet[f'AB{row_count}'].value = data_dict['annotation']
    sheet[f'AD{row_count}'].value = data_dict['author']
    sheet[f'AE{row_count}'].value = data_dict['publisher']
    sheet[f'AF{row_count}'].value = data_dict['publisher']
    sheet[f'AG{row_count}'].value = data_dict['year']
    sheet[f'AI{row_count}'].value = data_dict['series']
    sheet[f'AM{row_count}'].value = data_dict['weight']
    sheet[f'AN{row_count}'].value = data_dict['pages']
    sheet[f'AQ{row_count}'].value = data_dict['colored_pics']
    sheet[f'AS{row_count}'].value = data_dict['circulation']
    sheet[f'AT{row_count}'].value = data_dict['lang']
    sheet[f'AU{row_count}'].value = data_dict['orig_name']
    sheet[f'AX{row_count}'].value = data_dict['keeping']
    sheet[f'CY{row_count}'].value = data_dict['editor']
    sheet[f'DB{row_count}'].value = data_dict['illustrator']

    cells = sheet[row_count]

    if sheet_ranges != 'Без информации':
        sheet[f'G{row_count}'].value = 'Не облагается'

    if sheet_ranges == 'Букинистика' or sheet_ranges == 'Букинистика low':
        sheet[f'G{row_count}'].value = 'Букинистика'
    elif sheet_ranges == 'Second-hand' or sheet_ranges == 'Second-hand low':
        'Second-hand книги'

    necessary_column_list = ['D', 'K', 'L', 'M', 'N', 'O', 'V', 'X', 'AA']

    for column_letter in necessary_column_list:
        # print('Значение ячейки:', sheet[f'N{row_count}'].value)
        if sheet[f'{column_letter}{row_count}'].value == None or sheet[f'{column_letter}{row_count}'].value == '':
            # print('Empty')
            sheet[f'{column_letter}{row_count}'].fill = PatternFill(fill_type='solid', start_color='9400d3')

    if site_locators == BooksLocators:
        cells[0].fill = PatternFill(fill_type='solid', start_color='ff0000')

    work_book.save('Books_info_1.xlsx')


set_logger()
sys.excepthook = my_exception_hook

first_row_check = 1
COUNT = 0

COUNT_OF_BOOKS = 0

COUNT_ROW_BUKINISTICA_LOW_SHEET = 1
COUNT_ROW_SECOND_HAND_LOW_SHEET = 1

COUNT_LABIRINT_BOOKS = 0
COUNT_BOOKS_RU_BOOKS = 0

COUNT_ROW_BUKINISTICA_SHEET = 1
COUNT_ROW_SECOND_HAND_SHEET = 1
ROWS_COUNT_NO_YEAR = 1
ROWS_COUNT_NO_INFO = 1

time_interval_list = get_years_sequence(2011, 2025)

start_time = datetime.now()
requests.urllib3.disable_warnings(category=InsecureRequestWarning)

work_book = load_workbook('Books_info_1.xlsx')
sheet_ranges = work_book['Входные данные']

GENRES_LABIRINT_LIST = get_data_from_genres_file('Genres.xlsx')
GENRES_BOOKS_RU_LIST = get_data_from_genres_file('Genres_for_books_ru.xlsx')

user_agents_list = user_agents_unpack('user-agents.yml')

all_books_count = 0

# смотрим общее количество книг для логирования
for row in sheet_ranges.rows:
    all_books_count += 1

for row in sheet_ranges.rows:

    # TODO: сделать проверку на существование книги в excel
    user_agent = user_agents_list[r.randint(0, len(user_agents_list) - 1)]
    header = {
        'user-agent': user_agent,
    }

    ozon_exist = False
    COUNT += 1
    if first_row_check == 1:
        first_row_check += 1
        continue

    COUNT_OF_BOOKS += 1
    NUMBER, ISBN, KEEPING, LANG, BARCODE = get_input_data(row)

    site_locators, book_link = get_book_link(ISBN)

    TREE = get_site_tree(book_link)

    if not check_for_book_existing(ISBN):
        ROWS_COUNT_NO_INFO = filing_empty_sheet(ROWS_COUNT_NO_INFO, 'Без информации')
        logger.debug('Книги нет на сайтах')
        continue

    sheet_name, YEAR = define_which_sheet(time_interval_list, TREE)
    if sheet_name == 'Без информации':
        ROWS_COUNT_NO_INFO = filing_empty_sheet(ROWS_COUNT_NO_INFO, 'Без информации')
        continue

    book_ozon_page_link = get_link_to_ozon_page(ISBN)
    if book_ozon_page_link:
        ozon_tree = get_ozon_tree(book_ozon_page_link)
        ozon_exist = True
        photo_link = get_photo_link_from_ozon(ozon_tree)

        if photo_link:

            if not define_capture_size(photo_link):
                if sheet_name == 'Second-hand':
                    COUNT_ROW_SECOND_HAND_LOW_SHEET += 1
                    COUNT = COUNT_ROW_SECOND_HAND_LOW_SHEET
                    sheet_name = 'Second-hand low'

                elif sheet_name == 'Букинистика':
                    COUNT_ROW_BUKINISTICA_LOW_SHEET += 1
                    COUNT = COUNT_ROW_BUKINISTICA_LOW_SHEET
                    sheet_name = 'Букинистика low'

    else:
        photo_link = ''

    if sheet_name != 'Букинистика low' and sheet_name != 'Second-hand low':
        if site_locators == LabirintLocators:
            COUNT_LABIRINT_BOOKS += 1
        elif site_locators == BooksLocators:
            COUNT_BOOKS_RU_BOOKS += 1

    DATA_DICT = get_data(TREE)

    price_from_ozon = ''

    # нужно, чтоб проверить, все ли размеры есть
    if ozon_exist or site_locators == LabirintLocators:

        if DATA_DICT['height'] and DATA_DICT['width'] and DATA_DICT['length']:

            if not check_if_volume_is_more_than_one_and_more_than_one_and_a_half(DATA_DICT['length'],
                                                                                 DATA_DICT['width'],
                                                                                 DATA_DICT['height']):
                DATA_DICT['length'] = DATA_DICT['length'] * 0.96
                DATA_DICT['width'] = DATA_DICT['width'] * 0.96
                DATA_DICT['height'] = DATA_DICT['height'] * 0.94
            volume = DATA_DICT['length'] * DATA_DICT['width'] * DATA_DICT['height'] / 1000000
            minimal_price = calculate_minimal_price(volume)

            if ozon_exist:
                # TODO: цена с Ozon Premium на записываются в соответствующий столбец
                price_from_ozon = get_price_from_ozon(ozon_tree)
            else:
                price_from_ozon = ''

            if price_from_ozon:
                DATA_DICT['column_D_price'] = calculate_price_ozon_book_for_column_D(minimal_price, price_from_ozon)


            elif site_locators == LabirintLocators:

                # TODO: можно ли внести в словарь самую низкую цену Лабиринта на этапе парсинга?

                # если скидка есть, берем цену до скидки
                if DATA_DICT['old_price_with_discount_from_labirint']:
                    price_from_labirint = DATA_DICT['old_price_with_discount_from_labirint']

                    DATA_DICT['column_D_price'] = calculate_price_labirint_book(minimal_price,
                                                                                price_from_labirint)
                # если скидки нет, берем цену без скидки
                elif DATA_DICT['price_without_discount_from_labirint']:
                    price_from_labirint = DATA_DICT['price_without_discount_from_labirint']

                    DATA_DICT['column_D_price'] = calculate_price_labirint_book(minimal_price,
                                                                                price_from_labirint)
                # если цены в Лабиринте нет, оставляем столбец D пустым
            #     else:
            #         DATA_DICT['column_D_price'] = ''
            #
            # # если нет ни цены с озона, ни цены с лабиринта возвращаем пустое значение
            # else:
            #     DATA_DICT['column_D_price'] = ''

            # столбец F
            if DATA_DICT.get('column_D_price'):
                DATA_DICT['column_F_ozon_price'] = calculate_price_ozon_for_column_f(DATA_DICT['column_D_price'],
                                                                                     volume)

    # if not DATA_DICT.get('column_D_price'):
    #     DATA_DICT['column_F_ozon_price'] = ''


    # стоблец E
    if DATA_DICT.get('column_D_price'):
        DATA_DICT['column_E_price'] = calculate_old_price(price_from_ozon,
                                                          DATA_DICT['old_price_with_discount_from_labirint'],
                                                          DATA_DICT['price_without_discount_from_labirint'])
    # если столбец E пустой
    # else:
    #     DATA_DICT['column_E_price'] = None

    # TODO: заменить все '' на None
    # TODO: где возможно убрать пустуые строки, а в save_to_table заменить все методы на get

    if site_locators == LabirintLocators:
        site = 'Лабиринт'
    else:
        site = 'Books.ru'

    if sheet_name == 'Букинистика':
        COUNT_ROW_BUKINISTICA_SHEET += 1
        COUNT = COUNT_ROW_BUKINISTICA_SHEET

    elif sheet_name == 'Second-hand':
        COUNT_ROW_SECOND_HAND_SHEET += 1
        COUNT = COUNT_ROW_SECOND_HAND_SHEET

    save_to_table(COUNT, sheet_name, DATA_DICT)
    save_to_table(COUNT_OF_BOOKS + 1, 'Books', DATA_DICT)
    end_time = datetime.now()
    wasted_time = end_time - start_time
    logger.debug(f'Потрачено времени: {wasted_time}')
    logger.info(f'Книг обработано: {COUNT_OF_BOOKS}/{all_books_count}')

logger.info('Скрипт завершил работу!')
